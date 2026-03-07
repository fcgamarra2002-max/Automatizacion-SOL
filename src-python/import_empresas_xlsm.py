"""
Importar EMPRESAS.xlsm hacia la base SQLite operativa.
"""

import argparse
import os
import sqlite3

import openpyxl

from crypto import encrypt_clave, set_key_search_dir
from db_setup import setup_all


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def load_rows_from_excel(excel_path: str) -> list[dict]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = []
    for raw in worksheet.iter_rows(min_row=3, values_only=True):
        _, razon_social, ruc, usuario, clave, _, enabled = (list(raw) + [None] * 7)[:7]
        if not any(value not in (None, "") for value in (razon_social, ruc, usuario, clave)):
            continue
        if enabled is False:
            continue

        rows.append(
            {
                "RUC": normalize_text(ruc),
                "RazonSocial": normalize_text(razon_social),
                "UsuarioSOL": normalize_text(usuario),
                "ClaveSOL": normalize_text(clave),
                "Navegador": "chrome",
                "TipoPortal": "TRAMITES",
                "Motor": "selenium",
            }
        )

    return rows


def import_rows(db_path: str, rows: list[dict], replace: bool) -> tuple[int, int]:
    db_dir = os.path.dirname(os.path.abspath(db_path))
    set_key_search_dir(db_dir)

    if not setup_all(db_path):
        raise RuntimeError("No se pudo preparar la base SQLite antes de importar.")

    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        if replace:
            cursor.execute("DELETE FROM Empresas")
            cursor.execute("DELETE FROM sqlite_sequence WHERE name = 'Empresas'")

        inserted = 0
        updated = 0

        for row in rows:
            encrypted = encrypt_clave(row["ClaveSOL"])
            cursor.execute("SELECT Id FROM Empresas WHERE RUC = ?", (row["RUC"],))
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """
                    UPDATE Empresas
                    SET RazonSocial = ?, UsuarioSOL = ?, ClaveSOL = ?, Navegador = ?, TipoPortal = ?, Motor = ?
                    WHERE RUC = ?
                    """,
                    (
                        row["RazonSocial"],
                        row["UsuarioSOL"],
                        encrypted,
                        row["Navegador"],
                        row["TipoPortal"],
                        row["Motor"],
                        row["RUC"],
                    ),
                )
                updated += 1
            else:
                cursor.execute(
                    """
                    INSERT INTO Empresas (RUC, RazonSocial, UsuarioSOL, ClaveSOL, Navegador, TipoPortal, Motor)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row["RUC"],
                        row["RazonSocial"],
                        row["UsuarioSOL"],
                        encrypted,
                        row["Navegador"],
                        row["TipoPortal"],
                        row["Motor"],
                    ),
                )
                inserted += 1

        conn.commit()
        return inserted, updated
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="Importar empresas desde un archivo XLSM.")
    parser.add_argument("--excel", required=True, help="Ruta al archivo EMPRESAS.xlsm")
    parser.add_argument("--db", required=True, help="Ruta de salida para empresa.db")
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Reemplazar el contenido actual de la tabla Empresas antes de importar.",
    )
    args = parser.parse_args()

    rows = load_rows_from_excel(os.path.abspath(args.excel))
    inserted, updated = import_rows(os.path.abspath(args.db), rows, args.replace)
    print(
        f"Importacion completada. Filas fuente: {len(rows)}. Insertadas: {inserted}. Actualizadas: {updated}."
    )


if __name__ == "__main__":
    main()
